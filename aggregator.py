"""xGator, version 2.

Pulls chosen columns out of many Excel workbooks and writes every set
side by side into one copy of a template. v1 hardcoded columns G and H.
v2 takes a column spec typed by the user:

    G           one column
    G,H         several columns
    A-F         a range
    A,C-E,H     any mix
    ALL         every used column in each source file

The template is never opened for writing. The tool copies the template
to the chosen save path first and every write goes to the copy, so a
crash damages a throwaway file and never the original.
"""

import os
import platform
import re
import shutil
import subprocess
import sys

import openpyxl

# ---------------------------------------------------------------------------
# Defaults. Edit these two lines, or just use the fields in the window.
# ---------------------------------------------------------------------------
DEFAULT_COLUMN_SPEC = "G,H"     # what v1 always did
TEMPLATE_SHEET_INDEX = 1        # 0-based: 1 means the template's SECOND sheet


# ---------------------------------------------------------------------------
# Core logic. No GUI in this section, so every function here can be
# tested from a terminal without opening a window.
# ---------------------------------------------------------------------------

def column_letter_to_number(letter):
    """'A' -> 1, 'G' -> 7, 'AA' -> 27. Base-26 walk over the letters."""
    letter = letter.upper().strip()
    if not letter.isalpha():
        raise ValueError(f"'{letter}' is not a column letter")
    result = 0
    for char in letter:
        result = result * 26 + (ord(char) - ord("A") + 1)
    if result > 16384:
        raise ValueError(f"'{letter}' is past XFD, the last Excel column")
    return result


def parse_column_spec(spec):
    """Turn a typed spec into a sorted list of column numbers.

    'G'       -> [7]
    'G,H'     -> [7, 8]
    'A-C'     -> [1, 2, 3]
    'A,C-E,H' -> [1, 3, 4, 5, 8]
    'ALL'     -> [] and the empty list means every used column.

    Raises ValueError with a readable message on anything else, and the
    GUI shows that message instead of running.
    """
    spec = spec.strip()
    if not spec:
        raise ValueError("Type at least one column, a range like A-F, or ALL")
    if spec.upper() == "ALL":
        return []

    numbers = set()
    for part in spec.split(","):
        part = part.strip()
        if not part:
            continue
        m = re.fullmatch(r"([A-Za-z]{1,3})\s*-\s*([A-Za-z]{1,3})", part)
        if m:
            start = column_letter_to_number(m.group(1))
            end = column_letter_to_number(m.group(2))
            if start > end:
                raise ValueError(f"Range '{part}' runs backwards")
            numbers.update(range(start, end + 1))
        elif re.fullmatch(r"[A-Za-z]{1,3}", part):
            numbers.add(column_letter_to_number(part))
        else:
            raise ValueError(
                f"'{part}' is not a column letter, a range like A-F, or ALL")
    return sorted(numbers)


def describe_columns(numbers):
    """[7, 8] -> 'G, H' for the live label under the entry box."""
    if not numbers:
        return "every used column, per file"
    letters = []
    for n in numbers:
        name = ""
        while n > 0:
            n, rem = divmod(n - 1, 26)
            name = chr(ord("A") + rem) + name
        letters.append(name)
    return ", ".join(letters)


def aggregate(template_path, source_paths, save_path,
              column_numbers, sheet_index=TEMPLATE_SHEET_INDEX,
              log=print):
    """Copy the template to save_path, then write every source file's
    chosen columns into the copy, each file in the next free block of
    columns. Returns (files_processed, rows_written).

    column_numbers: sorted list from parse_column_spec. Empty list
    means ALL: each source file contributes its own used columns.
    """
    # Copy first. A missing template fails here, before any source
    # file is opened, and the original is never written to.
    shutil.copy(template_path, save_path)

    dest_wb = openpyxl.load_workbook(save_path)
    dest_ws = dest_wb.worksheets[sheet_index]

    current_dest_col = 1
    total_rows = 0

    for file_path in source_paths:
        # data_only=True reads the values Excel cached, not formula
        # text. Without it a formula cell lands as '=SUM(...)' and the
        # master file fills with text that looks like numbers.
        src_wb = openpyxl.load_workbook(file_path, data_only=True)
        src_ws = src_wb.active

        if column_numbers:
            cols = column_numbers
        else:
            # ALL: this file's own used width. Files with different
            # widths each contribute their own set.
            cols = list(range(1, src_ws.max_column + 1))

        min_col, max_col = min(cols), max(cols)
        wanted = {c - min_col for c in cols}       # offsets to keep

        rows_here = 0
        for i, row in enumerate(src_ws.iter_rows(min_col=min_col,
                                                 max_col=max_col,
                                                 values_only=True)):
            out_col = current_dest_col
            for offset, value in enumerate(row):
                if offset in wanted:
                    dest_ws.cell(row=1 + i, column=out_col, value=value)
                    out_col += 1
            rows_here += 1

        total_rows += rows_here
        current_dest_col += len(cols)              # next free block
        src_wb.close()
        log(f"{os.path.basename(file_path)}: {rows_here} rows, "
            f"{len(cols)} columns")

    dest_wb.save(save_path)
    dest_wb.close()
    return len(source_paths), total_rows


def resource_path(relative_path):
    """Find a bundled file both as a script and inside a PyInstaller
    one-file build, which unpacks itself into a temp folder named by
    sys._MEIPASS."""
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)


def open_file(filepath):
    """Open a file with the default application on any OS."""
    if platform.system() == "Windows":
        os.startfile(filepath)
    elif platform.system() == "Darwin":
        subprocess.call(["open", filepath])
    else:
        subprocess.call(["xdg-open", filepath])


# ---------------------------------------------------------------------------
# GUI. Same three steps and RUN button as v1, plus the column field.
# ---------------------------------------------------------------------------

def main():
    import tkinter as tk
    from tkinter import filedialog, messagebox

    state = {"template": "", "sources": [], "save": ""}

    root = tk.Tk()
    root.title("xGator")
    root.geometry("460x560")

    frame = tk.Frame(root, padx=10, pady=10)
    frame.pack(expand=True, fill=tk.BOTH)

    # -- Step 1: template --
    def select_template():
        path = filedialog.askopenfilename(
            title="1. Select the Template Master Excel File",
            filetypes=[("Excel Files", "*.xlsx")])
        if path:
            state["template"] = path
            lbl_template.config(text=f"Template: {os.path.basename(path)}")

    tk.Button(frame, text="1. Select Template File",
              command=select_template).pack(fill=tk.X, pady=5)
    lbl_template = tk.Label(frame, text="No template selected", fg="blue")
    lbl_template.pack()

    # -- Step 2: sources --
    def select_sources():
        paths = sorted(filedialog.askopenfilenames(
            title="2. Select the Source Data Files (one or many)",
            filetypes=[("Excel Files", "*.xlsx")]))
        if paths:
            state["sources"] = list(paths)
            lbl_sources.config(text=f"{len(paths)} source files selected")

    tk.Button(frame, text="2. Select Source Files",
              command=select_sources).pack(fill=tk.X, pady=(10, 5))
    lbl_sources = tk.Label(frame, text="No source files selected", fg="blue")
    lbl_sources.pack()

    # -- Step 3: columns --
    tk.Label(frame, text="3. Columns to extract from each source file:"
             ).pack(anchor="w", pady=(12, 0))
    entry_cols = tk.Entry(frame)
    entry_cols.insert(0, DEFAULT_COLUMN_SPEC)
    entry_cols.pack(fill=tk.X)
    lbl_cols = tk.Label(frame, text="", fg="blue")
    lbl_cols.pack()
    tk.Label(frame, fg="gray",
             text="One column G, a list G,H, a range A-F, a mix A,C-E,H,"
                  " or ALL").pack()

    def preview_columns(_event=None):
        """Reparse on every keystroke and show what will be pulled."""
        try:
            cols = parse_column_spec(entry_cols.get())
            lbl_cols.config(text=f"Will extract: {describe_columns(cols)}",
                            fg="blue")
        except ValueError as exc:
            lbl_cols.config(text=str(exc), fg="red")

    entry_cols.bind("<KeyRelease>", preview_columns)
    preview_columns()

    # -- Step 4: save location --
    def select_save():
        path = filedialog.asksaveasfilename(
            title="4. Select Save Location for the New Master File",
            filetypes=[("Excel Files", "*.xlsx")],
            defaultextension=".xlsx")
        if path:
            state["save"] = path
            lbl_save.config(text=f"Save as: {os.path.basename(path)}")

    tk.Button(frame, text="4. Select Save Location",
              command=select_save).pack(fill=tk.X, pady=(12, 5))
    lbl_save = tk.Label(frame, text="No save location selected", fg="blue")
    lbl_save.pack()

    # -- RUN --
    def run():
        if not state["template"] or not state["sources"] or not state["save"]:
            messagebox.showerror(
                "Error", "Complete steps 1, 2 and 4 before running.")
            return
        try:
            cols = parse_column_spec(entry_cols.get())
        except ValueError as exc:
            messagebox.showerror("Column spec", str(exc))
            return

        lbl_run.config(text="Running...")
        root.update()
        try:
            files, rows = aggregate(state["template"], state["sources"],
                                    state["save"], cols)
        except Exception as exc:
            messagebox.showerror("Error", f"An error occurred:\n{exc}")
            lbl_run.config(text="Error. Try again.")
            return

        lbl_run.config(text=f"DONE. {files} files, {rows} rows.")
        if messagebox.askyesno(
                "Done",
                f"Master file created.\n\n"
                f"{files} source files processed\n"
                f"{rows} total rows written\n"
                f"Columns: {describe_columns(cols)}\n\n"
                f"Open the file now?"):
            open_file(state["save"])

    tk.Button(frame, text="RUN", font=("Arial", 14, "bold"),
              bg="green", fg="white", command=run
              ).pack(fill=tk.X, pady=20, ipady=10)
    lbl_run = tk.Label(frame, text="Ready", font=("Arial", 12))
    lbl_run.pack()

    root.mainloop()


if __name__ == "__main__":
    main()
